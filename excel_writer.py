import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import numbers


def month_span(dt: datetime) -> tuple[str, str]:
    month_name = dt.strftime("%B")
    month_name_es = {
        "January": "Enero",
        "February": "Febrero",
        "March": "Marzo",
        "April": "Abril",
        "May": "Mayo",
        "June": "Junio",
        "July": "Julio",
        "August": "Agosto",
        "September": "Septiembre",
        "October": "Octubre",
        "November": "Noviembre",
        "December": "Diciembre",
    }[month_name]
    year = dt.strftime("%Y")
    return month_name_es, year


def get_excel_path(base_directory: str, dt: datetime) -> str:
    month_name_es, year = month_span(dt)
    os.makedirs(base_directory, exist_ok=True)
    file_name = f"LecturasMuroContencion{month_name_es}{year}.xlsx"
    return os.path.join(base_directory, file_name)


def _ensure_headers(ws: Worksheet) -> None:
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(["Fecha", "Dataloger", "Nombre", "Lectura"])


def append_rows(excel_path: str, rows: list[dict]) -> None:
    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Lecturas"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lecturas"

    _ensure_headers(ws)

    for row in rows:
        ws.append([
            row.get("Fecha", ""),
            row.get("Dataloger", ""),
            row.get("Nombre", ""),
            row.get("Lectura", ""),
        ])
        last = ws.max_row
        # Formatos: fecha y número
        ws.cell(row=last, column=1).number_format = "yyyy/mm/dd hh:mm"
        # Dataloger como número entero si es posible
        try:
            datalog = row.get("Dataloger", "")
            if isinstance(datalog, str) and datalog.isdigit():
                ws.cell(row=last, column=2).value = int(datalog)
            ws.cell(row=last, column=2).number_format = numbers.FORMAT_NUMBER
        except Exception:
            ws.cell(row=last, column=2).number_format = numbers.FORMAT_NUMBER
        try:
            # Asegurar tipo float para que Excel lo trate como número
            val = row.get("Lectura", "")
            if isinstance(val, str):
                # Si viene texto, intentar convertir reemplazando coma por punto
                val_float = float(val.replace(".", "").replace(",", ".")) if "," in val else float(val)
                ws.cell(row=last, column=4).value = val_float
            ws.cell(row=last, column=4).number_format = numbers.FORMAT_NUMBER_00
        except Exception:
            ws.cell(row=last, column=4).number_format = numbers.FORMAT_NUMBER_00

    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)
    wb.save(excel_path)


